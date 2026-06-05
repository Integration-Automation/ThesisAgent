"""Excel (.xlsx) exporter — one sheet, one row per paper.

The workbook is structured for human use: header row in bold, top row frozen,
column widths sized to the content (capped so long abstracts don't blow out
the viewport), abstract cells wrap text. Hyperlinks are clickable in
Excel / Numbers / LibreOffice.
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from thesisagents.core.constants import EXPORT_XLSX
from thesisagents.core.exceptions import ExportError
from thesisagents.core.models import ExportOptions, Paper, PaperCollection
from thesisagents.exporters.base import Exporter

_COLUMNS: tuple[tuple[str, int], ...] = (
    ("#", 4),
    ("Title", 60),
    ("Authors", 36),
    ("Year", 6),
    ("Source", 28),
    ("Indexed via", 12),
    ("DOI", 28),
    ("URL", 40),
    ("PDF", 40),
    ("Citations", 10),
    ("Abstract", 80),
)


class XlsxExporter(Exporter):
    format = EXPORT_XLSX
    extension = "xlsx"

    def export(self, collection: PaperCollection, options: ExportOptions) -> Path:
        try:
            workbook = self._build(collection, options.include_abstract)
        except Exception as err:
            raise ExportError(self.format, f"render failed: {err}") from err
        out_path = self.resolve_out_path(collection, options)
        workbook.save(str(out_path))
        return out_path

    def _build(self, collection: PaperCollection, include_abstract: bool) -> Workbook:
        workbook = Workbook()
        self._write_papers_sheet(workbook, collection, include_abstract)
        self._write_meta_sheet(workbook, collection)
        return workbook

    @staticmethod
    def _write_papers_sheet(
        workbook: Workbook, collection: PaperCollection, include_abstract: bool
    ) -> None:
        sheet = workbook.active
        sheet.title = "Papers"
        _write_header(sheet)
        for row_index, paper in enumerate(collection.papers, start=2):
            _write_paper_row(
                sheet, row_index, row_index - 1, paper, include_abstract=include_abstract
            )
        _apply_column_widths(sheet)
        sheet.freeze_panes = "A2"

    @staticmethod
    def _write_meta_sheet(workbook: Workbook, collection: PaperCollection) -> None:
        sheet = workbook.create_sheet("Query")
        rows = [
            ("Query keywords", collection.query.keywords),
            ("Sources", ", ".join(collection.query.sources)),
            ("Max per source", collection.query.max_results),
            ("Year from", collection.query.year_from or ""),
            ("Year to", collection.query.year_to or ""),
            ("Min citations", collection.query.min_citations or ""),
            ("Result count", len(collection)),
            ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ]
        bold = Font(bold=True)
        for index, (key, value) in enumerate(rows, start=1):
            sheet.cell(row=index, column=1, value=key).font = bold
            sheet.cell(row=index, column=2, value=value)
        sheet.column_dimensions["A"].width = 22
        sheet.column_dimensions["B"].width = 60


def _write_header(sheet) -> None:
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    centre = Alignment(horizontal="center", vertical="center")
    for col_index, (label, _width) in enumerate(_COLUMNS, start=1):
        cell = sheet.cell(row=1, column=col_index, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = centre
    sheet.row_dimensions[1].height = 22


def _write_paper_row(
    sheet, row: int, index: int, paper: Paper, *, include_abstract: bool
) -> None:
    sheet.cell(row=row, column=1, value=index)
    sheet.cell(row=row, column=2, value=paper.title)
    sheet.cell(row=row, column=3, value=", ".join(paper.authors))
    sheet.cell(row=row, column=4, value=paper.year)
    sheet.cell(row=row, column=5, value=paper.venue or "")
    sheet.cell(row=row, column=6, value=paper.source)
    sheet.cell(row=row, column=7, value=paper.doi or "")
    _write_hyperlink(sheet.cell(row=row, column=8, value=paper.url), paper.url)
    if paper.pdf_url:
        _write_hyperlink(sheet.cell(row=row, column=9, value=paper.pdf_url), paper.pdf_url)
    sheet.cell(row=row, column=10, value=paper.citation_count)
    if include_abstract:
        sheet.cell(row=row, column=11, value=paper.abstract)
    # Wrap the long-text columns.
    wrap = Alignment(wrap_text=True, vertical="top")
    for col in (2, 3, 11):
        sheet.cell(row=row, column=col).alignment = wrap


def _write_hyperlink(cell, url: str) -> None:
    if not url:
        return
    cell.hyperlink = url
    cell.font = Font(color="0563C1", underline="single")


def _apply_column_widths(sheet) -> None:
    for col_index, (_label, width) in enumerate(_COLUMNS, start=1):
        sheet.column_dimensions[get_column_letter(col_index)].width = width
