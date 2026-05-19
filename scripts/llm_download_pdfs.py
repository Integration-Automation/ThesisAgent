"""Batch LLM-driven PDF download from an aggregate xlsx.

Reads a "Papers" sheet produced by ``XlsxExporter`` (columns
``# | Title | Authors | Year | Source | Indexed via | DOI | URL |
PDF | Citations | Abstract``), groups rows by publisher, opens a
SINGLE visible Chrome session, and walks each paper in turn. Reusing
one driver across N papers means: cookies / VPN auth survive, captcha
(if any) is solved once, and the per-paper overhead drops to one
``driver.get`` + a download wait.

Supported publishers (URL host → handler):
* ``ieeexplore.ieee.org`` → ``download_ieee`` (uses arnumber from URL)
* ``dl.acm.org`` → ``download_acm`` (uses DOI)
* ``link.springer.com`` → ``download_springer`` (uses DOI)

Rows whose URL host isn't in the table above are skipped with a note.

Usage:
    .venv\\Scripts\\python.exe -m scripts.llm_download_pdfs <xlsx_path>
    .venv\\Scripts\\python.exe -m scripts.llm_download_pdfs <xlsx_path> \\
        --publishers ieee,acm

Outputs land in ``exports/_llm_scratch/pdfs/`` next to the per-paper
downloads from the single-paper CLIs. Exit code: 0 when every paper
downloaded, 1 when one or more failed.
"""

from __future__ import annotations

import argparse
import contextlib
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from autopapertoppt.fetchers import webrunner_browser
from scripts._pdf_downloaders import (
    dispatch_for_url,
    download_aclanthology,
    download_acm,
    download_arxiv,
    download_ieee,
    download_neurips,
    download_openreview,
    download_springer,
)

OUT_DIR = Path(r"D:\Codes\AutoPaperToPPT\exports\_llm_scratch\pdfs")
OUT_DIR.mkdir(parents=True, exist_ok=True)


def _invoke(publisher: str, args: tuple, driver: Any, out_dir: Path) -> Path | None:
    """Route ``(publisher, *args)`` to its downloader."""
    if publisher == "ieee":
        return download_ieee(driver, args[0], out_dir)
    if publisher == "acm":
        return download_acm(driver, args[0], out_dir)
    if publisher == "springer":
        return download_springer(driver, args[0], out_dir)
    if publisher == "arxiv":
        return download_arxiv(driver, args[0], out_dir)
    if publisher == "acl":
        return download_aclanthology(driver, args[0], out_dir)
    if publisher == "neurips":
        return download_neurips(driver, args[0], args[1], out_dir)
    if publisher == "openreview":
        return download_openreview(driver, args[0], out_dir)
    raise ValueError(f"unknown publisher: {publisher!r}")


def _load_papers(xlsx_path: Path) -> list[dict[str, str]]:
    """Return one dict per row in the 'Papers' sheet."""
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    sheet = wb["Papers"]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h or "").strip() for h in rows[0]]
    out: list[dict[str, str]] = []
    for row in rows[1:]:
        record: dict[str, str] = {}
        for header, value in zip(headers, row, strict=False):
            record[header] = "" if value is None else str(value)
        out.append(record)
    return out


def _plan(
    papers: list[dict[str, str]], publisher_filter: set[str],
) -> list[tuple[str, tuple[str, ...], str]]:
    """Pick rows we can download. Returns ``(publisher, identifier_parts, title_preview)``.

    ``identifier_parts`` is a tuple so multi-part identifiers (NeurIPS uses
    ``(year, hash)``) ride through the planner without special casing.
    """
    plan: list[tuple[str, tuple[str, ...], str]] = []
    seen: set[tuple] = set()
    for row in papers:
        url = row.get("URL", "") or row.get("Url", "")
        doi = row.get("DOI", "") or row.get("Doi", "")
        dispatch = dispatch_for_url(url, doi or None)
        if dispatch is None:
            continue
        publisher, *identifier_parts = dispatch
        ident_tuple = tuple(identifier_parts)
        if publisher_filter and publisher not in publisher_filter:
            continue
        key = (publisher, *ident_tuple)
        if key in seen:
            continue
        seen.add(key)
        title = (row.get("Title") or "")[:60]
        plan.append((publisher, ident_tuple, title))
    return plan


def _run(xlsx_path: Path, publisher_filter: set[str]) -> int:
    papers = _load_papers(xlsx_path)
    plan = _plan(papers, publisher_filter)
    if not plan:
        print(f"[plan] nothing to download from {xlsx_path}", flush=True)
        return 0

    by_pub: dict[str, list[tuple[str, str]]] = {}
    for publisher, ident_tuple, title in plan:
        by_pub.setdefault(publisher, []).append(("/".join(ident_tuple), title))
    print(
        "[plan] " + ", ".join(
            f"{p}={len(rows)}" for p, rows in sorted(by_pub.items())
        ),
        flush=True,
    )

    print(f"[boot] visible Chrome with download_dir={OUT_DIR}", flush=True)
    driver = webrunner_browser.make_driver(download_dir=str(OUT_DIR))
    failures: list[tuple[str, str, str]] = []
    successes: list[tuple[str, str, Path]] = []
    try:
        for publisher, ident_tuple, title in plan:
            ident_str = "/".join(ident_tuple)
            print(
                f"\n=== {publisher} :: {ident_str} :: {title!r} ===",
                flush=True,
            )
            try:
                saved = _invoke(publisher, ident_tuple, driver, OUT_DIR)
            except Exception as err:  # noqa: BLE001 — selenium raises many types
                print(f"[err]  {publisher} {ident_str} raised: {err}", flush=True)
                failures.append((publisher, ident_str, f"exception: {err}"))
                continue
            if saved is None:
                failures.append((publisher, ident_str, "no PDF produced"))
            else:
                successes.append((publisher, ident_str, saved))
    finally:
        with contextlib.suppress(Exception):
            driver.quit()

    print("\n=== summary ===")
    print(f"  ok={len(successes)}  fail={len(failures)}  total={len(plan)}")
    for pub, ident, path in successes:
        print(f"  [ok]   {pub} {ident} -> {path.name} ({path.stat().st_size:,} bytes)")
    for pub, ident, reason in failures:
        print(f"  [fail] {pub} {ident} :: {reason}")
    return 0 if not failures else 1


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__.split("\n")[0])
    parser.add_argument("xlsx", help="path to an XlsxExporter 'Papers' sheet")
    parser.add_argument(
        "--publishers",
        default="",
        help="comma-separated subset of {ieee,acm,springer}; default = all",
    )
    args = parser.parse_args()
    xlsx_path = Path(args.xlsx)
    if not xlsx_path.is_file():
        print(f"[err] xlsx not found: {xlsx_path}")
        return 2
    publisher_filter = {p.strip() for p in args.publishers.split(",") if p.strip()}
    return _run(xlsx_path, publisher_filter)


if __name__ == "__main__":
    sys.exit(main())
