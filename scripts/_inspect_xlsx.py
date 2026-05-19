"""Throwaway: print every row of a Papers sheet for the LLM to inspect."""
import sys

from openpyxl import load_workbook

wb = load_workbook(sys.argv[1], read_only=True, data_only=True)
ws = wb["Papers"]
rows = list(ws.iter_rows(values_only=True))
hdr = rows[0]
title_i = hdr.index("Title")
via_i = hdr.index("Indexed via")
doi_i = hdr.index("DOI")
url_i = hdr.index("URL")
for i, r in enumerate(rows[1:], 1):
    title = (r[title_i] or "")[:60]
    print(f"[{i}] via={r[via_i]:8} | {title}")
    print(f"    URL={r[url_i]}")
    print(f"    DOI={r[doi_i]}")
